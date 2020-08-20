import bs4 as bs
import configparser
import datetime as dt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import pandas as pd
import pandas_datareader.data as pdd
from PyQt5.QtWidgets import QApplication, QSystemTrayIcon, QMenu
from PyQt5.QtGui import QIcon
import requests
import sys
import threading
import time
from win10toast import ToastNotifier


def check_for_warning(name, price, sheet, condition=None, threshold=None, direction=None, sl=None, tp=None):
    toaster = ToastNotifier()
    trigger = False

    if sheet == 'Alert':
        if condition == '>' and price >= threshold:
            trigger = True
        elif condition == '<' and price <= threshold:
            trigger = True
        message = str(name) + ' triggered an Alert! Price: ' + str(price)

    elif sheet == 'SLTP':
        if direction == 'Long':
            if price <= sl:
                trigger = True
                sltp_type = 'Stop-Loss'
            elif price >= tp:
                trigger = True
                sltp_type = 'Take-Profit'
        elif direction == 'Short':
            if price >= sl:
                trigger = True
                sltp_type = 'Stop-Loss'
            elif price <= tp:
                trigger = True
                sltp_type = 'Take-Profit'

        if trigger:
            message = str(name) + ' triggered a' + str(sltp_type) + ' Alert! Price: ' + str(price)

    elif sheet == 'Options':
        if direction == 'call':
            if price <= sl:
                trigger = True
        elif direction == 'put':
            if price >= sl:
                trigger = True

        message = 'Option ' + str(name) + ' triggered a SL-Alert! Price: ' + str(price)

    if trigger:
        toaster.show_toast('Finance', message, duration=999999, icon_path='stocks.ico', threaded=False)
        url = 'https://www.pushsafer.com/api?k=vLnoDWjGWkYYYPiwBqaH&d=19518&i=9&c=%239900ff&v=1&pr=2&t=' + 'Price Alert' + '&m=' + str(message)
        requests.get(url)


def save_sheet(wb):
    try:
        wb.save(os.getcwd() + '\\Data.xlsx')
    except PermissionError:
        print(str(dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")) + ' File could not be saved!')


def update_data(sleep_time=600, show_tray=True, local_timezone='Europe/Berlin'):

    if show_tray:
        thread = threading.Thread(target=tray_icon_handler, name='Price Alert Tray')
        thread.start()



    sheets = ['Alert', 'SLTP', 'Change', 'Options', 'Calendar']
    change_times = {'1 week': 7, '1 month': 30, '1 quarter': 90, '1 year': 360}

    checking = True
    while checking:
        start = dt.datetime(dt.datetime.now().year - 1, 1, 1)
        end = dt.datetime.now()

        for name in sheets:
            try:
                wb = openpyxl.load_workbook(os.getcwd() + '\\Data.xlsx', data_only=True)
                sheet = wb[name]
                update_time = []

                if name in ['Alert', 'SLTP', 'Change', 'Options']:
                    data_rows = []

                    for row in sheet['A9':'G' + str(sheet.max_row)]:
                        data_cols = []
                        for cell in row:
                            data_cols.append(cell.value)
                        data_rows.append(data_cols)

                    df = pd.DataFrame(data_rows[1:], columns=data_rows[0])

                    if name in ['Alert', 'SLTP']:
                        prices = []
                        for row in df.iterrows():
                            request = pdd.DataReader(row[1]['Symbol'], 'yahoo', start, end)
                            prices.append(round(request['Close'][-1], 2))
                            update_time.append(dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

                            if name == 'Alert':
                                check_for_warning(name=row[1]['Name'], price=round(request['Close'][-1], 2), sheet=name, condition=row[1]['Condition'], threshold=row[1]['Threshold'])
                            elif name == 'SLTP':
                                check_for_warning(name=row[1]['Name'], price=round(request['Close'][-1], 2), sheet=name, direction=row[1]['Direction'], sl=row[1]['SL'], tp=row[1]['TP'])

                        df['Last Update'] = update_time
                        df['Price'] = prices

                    elif name in ['Change']:
                        prices = [[], [], [], [], []]
                        for row in df.iterrows():
                            request = pdd.DataReader(row[1]['Symbol'], 'yahoo', start, end)
                            prices[4].append(round(request['Close'][-1], 2))
                            request.reset_index(inplace=True)
                            update_time.append(dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                            for chg_time in change_times:
                                search_date = dt.datetime.now().date() - dt.timedelta(days=change_times[chg_time])
                                for req_row in request.iterrows():
                                    if req_row[1]['Date'] >= search_date:
                                        prices[list(change_times).index(chg_time)].append(round(req_row[1]['Close'], 2))
                                        break

                        df['Last Update'] = update_time
                        df['1 week'] = prices[0]
                        df['1 month'] = prices[1]
                        df['1 quarter'] = prices[2]
                        df['1 year'] = prices[3]
                        df['Price'] = prices[4]

                    elif name == 'Options':
                        prices = []
                        for row in df.iterrows():
                            wkn = row[1]['Symbol'][-7:-1]
                            url = 'https://www.ariva.de/optionsscheine/' + wkn + '/historische_kurse?boerse_id=47'

                            resp = requests.get(url)
                            html = bs.BeautifulSoup(resp.text, 'lxml')

                            price = bs.BeautifulSoup(str(html.select('div.snapshotQuotesBox.designZERT')), 'lxml').select('td.first')[1].text.split()[0]

                            price = float(str(price).replace(',', '.'))

                            prices.append(round(price, 2))
                            update_time.append(dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

                            check_for_warning(name=row[1]['Name'], price=price, sheet=name, condition=row[1]['Type'], threshold=row[1]['SL'])

                        df['Last Update'] = update_time
                        df['Price'] = prices

                    rows = dataframe_to_rows(df, index=False)
                    for r_idx, row in enumerate(rows, 9):
                        for c_idx, value in enumerate(row, 1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)

                    save_sheet(wb)

                elif name in ['Calendar']:
                    data_rows = []
                    earnings_table_full = pd.DataFrame()

                    for row in sheet['A9':'B' + str(sheet.max_row)]:
                        data_cols = []
                        for cell in row:
                            data_cols.append(cell.value)
                        data_rows.append(data_cols)

                    df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
                    df.dropna(inplace=True)

                    for row in df.iterrows():
                        url = 'https://finance.yahoo.com/calendar/earnings/?symbol=' + row[1]['Symbol']

                        resp = requests.get(url)
                        calendar_source = bs.BeautifulSoup(resp.text, 'html5lib')
                        calendar_table = calendar_source.select('table', {'class': "W(100%)"})

                        stock_df = pd.DataFrame(pd.read_html(str(calendar_table))[0])
                        stock_df['Earnings Date'] = pd.to_datetime(stock_df['Earnings Date'], format='%b %d, %Y, %I %p%Z')
                        stock_df['Earnings Date'] = stock_df['Earnings Date'].dt.tz_convert(tz=local_timezone).dt.strftime('%d.%m.%Y %H:%M')

                        earnings_table_full = pd.concat([earnings_table_full, stock_df], ignore_index=True)

                    earnings_table_full.drop(columns='Surprise(%)', inplace=True)

                    for i, row in earnings_table_full.iterrows():
                        old = dt.datetime.today() - dt.timedelta(days=7)
                        new = dt.datetime.today() + dt.timedelta(days=360)
                        if not old <= dt.datetime.strptime(row['Earnings Date'], '%d.%m.%Y %H:%M') <= new:
                            earnings_table_full.drop(index=i, inplace=True)

                    earnings_table_full['Last Update'] = [dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S") for x in range(len(earnings_table_full['Earnings Date']))]

                    earnings_table_full.fillna(value='', inplace=True)
                    earnings_table_full['Earnings Date'] = pd.to_datetime(earnings_table_full['Earnings Date'], format='%d.%m.%Y %H:%M')
                    earnings_table_full.sort_values(by='Earnings Date', ascending=True, inplace=True)

                    rows = dataframe_to_rows(earnings_table_full, index=False)
                    for r_idx, row in enumerate(rows, 9):
                        for c_idx, value in enumerate(row, 5):
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                    save_sheet(wb)

            except:
                last_update.setText('ERROR OCCURRED! Name: ' + str(name) + '  Time: ' + str(dt.datetime.now().strftime("%H:%M")))
                continue

        if show_tray:
            last_update.setText('Last update: ' + str(dt.datetime.now().strftime("%H:%M")))

        print(str(dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")) + ' Update done!')
        time.sleep(sleep_time)
        # app.exit()


def tray_icon_handler():

    app = QApplication(sys.argv)

    tray_icon = QSystemTrayIcon(QIcon('stocks.ico'), parent=app)
    tray_icon.setToolTip('Price Alert Running!')
    tray_icon.show()

    menu = QMenu()
    global last_update
    last_update = menu.addAction('Last update: - - - ')

    open_excel = menu.addAction('Open Price Alert Excel Sheet')
    open_excel.triggered.connect(lambda: os.startfile(os.getcwd() + '\\Price_Alert.xlsm'))

    exit_action = menu.addAction('Exit')
    exit_action.triggered.connect(lambda: os._exit(0))

    tray_icon.setContextMenu(menu)

    sys.exit(app.exec_())


def startup():

    config = configparser.ConfigParser()
    config.read(os.getcwd() + '\\' + 'config.ini')

    sleep_time = config['General Settings']['sleep_time']
    show_tray = config['General Settings'].getboolean('show_tray')
    local_timezone = config['Calendar']['local_timezone']


    update_data(sleep_time, show_tray, local_timezone)


startup()
